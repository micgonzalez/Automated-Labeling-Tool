# This version two of auto labeling tool

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Replacement numbers for year column
replacement_pair = {'2018â€“': '2018-','2011â€“2016': '2011-2016', '2020â€“': '2020-', 
                    '2015â€“2016': '2015-2016','2017â€“': '2017-', '1993â€“1994': '1993-1994', 
                    '1993â€“2000': '1993-2000','2011â€“': '2011-', '1988â€“1990': '1988-1990', 
                    '2012â€“2014': '2012-2014', '1991â€“1992': '1991-1992', '2019â€“': '2019-', 
                    '2016â€“2019': '2016-2019','2015â€“': '2015-', '2014â€“': '2014-', '2010â€“': '2010-',
                    '2012â€“2016': '2012-2016','2014â€“2015': '2014-2015', '2011â€“2016': '2011-2016', 
                    '2009â€“2010': '2009-2010','2015â€“2018': '2015-2018', '2006â€“2016': '2006-2016', 
                    '2016â€“': '2016-','2012â€“2015': '2012-2015', '2017â€“': '2017-','2015â€“2019': '2015-2019',
                    '2016â€“2018': '2016-2018', '1985â€“1991': '1985-1991','1996â€“1999': '1996-1999', 
                    '1987â€“1990': '1987-1990', '1998â€“1999': '1998-1999','2002â€“2007': '2002-2007', 
                    '2005â€“2009': '2005-2009', '1996â€“1997': '1996-1997','1997â€“2001': '1997-2001', 
                    '2006â€“2008': '2006-2008', '1992â€“1994': '1992-1994','2012â€“': '2012-', 
                    '2004â€“2016': '2004-2016', '2000â€“2003': '2000-2003','1994â€“1996': '1994-1996', 
                    '2006â€“2010': '2006-2010', '2010â€“2014': '2010-2014','2014â€“2017': '2014-2017',
                    '1992â€“1993': '1992-1993', '2006â€“2013': '2006-2013','2006â€“2011': '2006-2011', 
                    '2013â€“': '2013-', '2008â€“': '2008-', '2011â€“2015': '2011-2015','2003â€“2006': '2003-2006',
                    '2001â€“2004': '2001-2004', '1994â€“1998': '1994-1998','1996â€“1998': '1996-1998', 
                    '1992â€“1997': '1992-1997', '2012â€“2017': '2012-2017','2015â€“2016': '2015-2016', 
                    '2017â€“2019': '2017-2019', '1955â€“1958': '1955-1958','2007â€“2010': '2007-2010', 
                    '1998â€“2004': '1998-2004', '2004â€“2006': '2004-2006','2007â€“2015': '2007-2015',
                    '2010â€“2013': '2010-2013', '1997â€“1999': '1997-1999', '1999â€“2001': '1999-2001',
                    '1981â€“1982': '1981-1982', '1981â€“1986': '1981-1986', '1999â€“2005': '1999-2005',
                    '1979â€“1980': '1979-1980', '2014â€“2018': '2014-2018', '2008â€“2020': '2008-2020',
                    '1990â€“1991': '1990-1991', '2017â€“2020': '2017-2020', '2000â€“2005': '2000-2005',
                    '2003â€“2007': '2003-2007', '1955â€“': '1955-', '2010â€“2012': '2010-2012',
                    '1999â€“2000': '1999-2000', '2001â€“2004': '2001-2004', '2015â€“2016': '2015-2016',
                    '1988â€“1991': '1988-1991', '2001â€“2005': '2001-2005', '2006â€“2009': '2006-2009',
                    '2005â€“2008': '2005-2008', '2008â€“2011': '2008-2011', '2009â€“2011': '2009-2011',
                    '1995â€“1999': '1995-1999', '2007â€“2012': '2007-2012', '2008â€“2009': '2008-2009',
                    '2000â€“2003': '2000-2003', '1989â€“': '1989-', '2013â€“2015': '2013-2015', 
                    '2008â€“2012': '2008-2012', '2012â€“2013': '2012-2013'}


# Define yellow fill color for updated cells
corrected_cells = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")

wb = load_workbook("disney_plus_shows_workbook.xlsx")
ws = wb['disney_plus_shows']
ws['T1'] = "Good Review"
ws['U1'] = "Bad Review"

not_available = ["N/A", "None"]

approv_list = ["Bearly Asleep", "Beezy Bear", "Canine Caddy",
               "Donalds Dog Laundry", "Donalds Golf Game",
               "Donalds Tire Trouble", "Double Dribble",
               "Dragon Around", "How to Fish", "How to Play Baseball",
               "How to Swim", "Lets Stick Together", "Lonesome Ghosts",
               "Mr. Mouse Takes a Trip", "On Ice", "Out of Scale",
               "Plutos Purchase", "Santas Workshop", "Sea Scouts",
               "Society Dog Show", "Swiss Family Robinson",
               "The Golden Touch", "The New Neighbor", "The Pied Piper",
               "The Wise Little Hen", "Trailer Horn", "Tugboat Mickey"]

g_list = ["Diving with Dolphins", "Elephant", "Geris Game",
          "The Boy Who Talked to Badgers"]


pg_list = ["Into the Canyon", "Into the Okavango"]


tvg_list = ["Be Our Chef", "Boat Builders", "Bug Juice: My Adventures at Camp",
            "Family Sundays", "Fast Layne", "Disneyland Around the Seasons",
            "What is Cheese?", "What is Reading?", "Fuzzbucket",
            "Kingdom of the Blue Whale", "Lamp Life", "My Dog, the Thief: Part 1",
            "Shop Class", "The Ghosts of Buxley Hall",
            "The Imagineering Story", "The Lodge", "Unlikely Animal Friends",
            "Wind", "BarkFest"]

tvpg_list = ["Auntie Edna", "Continent 7: Antarctica", "Diana: In Her Own Words",
             "The Evermoor Chronicles", "Dog: Impossible", "Dr. Oakley, Yukon Vet",
             "Dr. T, Lone Star Vet", "Drain the Bermuda Triangle", "Drain the Titanic",
             "Earth Live", "Easter Island Underworld", "Empire of Dreams: The Story of the Star Wars Trilogy",
             "Expedition Amelia", "Expedition Mars", "Disneys Fairy Tale Weddings",
             "Fantastic Four: The Animated Series", "Fantastic Four: Worlds Greatest Heroes",
             "Hostile Planet", "How Dogs Got Their Shapes", "Incredible! The Story of Dr. Pol",
             "Nature Untamed", "Kingdom of the Blue Whale",
             "Kingdom of the White Wolf", "Lost Cities with Albert Lin", "Lost Treasures of the Maya",
             "Man Among Cheetahs", "MARS: Inside SpaceX", "Marvel 75 Years: From Pulp to Pop!",
             "Marvel Rising: Battle of the Bands", "Marvel Rising: Heart of Iron", "Marvel Rising: Initiation",
             "Marvel Rising: Secret Warriors", "Marvel Super Hero Adventures", "Rocket & Groot",
             "Miracle at Midnight", "Miracle Landing on the Hudson", "Mission to the Sun",
             "One Day at Disney", "Origins: The Journey of Humankind", "Out There with Jack Randall",
             "Paris to Pittsburgh", "Pick of the Litter", "Primal Survivor",
             "Rocky Mountain Animal Rescue", "Ruby Bridges", "Sea of Hope: Americas Underwater Treasures",
             "The Secret of Christs Tomb", "Secrets of the King Cobra", "Secrets of the Zoo",
             "Sharks of Lost Island", "Soy Luna", "Supercar Superbuild",
             "The Flood", "The Incredible Dr. Pol", "Blue Ribbon Kids",
             "The Legend of Mordu", "The Lodge", "The Lost Tomb of Alexander the Great",
             "Titanic: 20 Years Later with James Cameron", "Tutankhamens Treasures", "Wind",
             "Winged Seduction: Birds of Paradise", "Year Million"]

tvy_list = ["Disneys Doug","Mission Force One", "Muppet Moments"]


tvy7_list = ["Billy Dilleys Super-Duper Subterranean Summer",
             "Crash & Bernstein", "Descendants: Wicked World",
             "Imagination Movers", "Lego Star Wars: Droid Tales",
             "LEGO Star Wars: The Resistance Rises", "The New Yoda Chronicles: Clash of the Skywalkers",
             "The New Yoda Chronicles: Raid on Coruscant", "The New Yoda Chronicles: Escape from the Jedi Temple",
             "Lego Star Wars: The Yoda Chronicles", "Marvel Rising: Chasing Ghosts",
             "Ant-Man", "Spider-Man", "Star Wars: Forces of Destiny", "Star Wars Blips"]

tv14_list = ["Insider", "Drain Alcatraz", "Drain the Sunken Pirate City",
             "Gordon Ramsay: Uncharted", "Viking Warrior Women"]

nr_list = ["A Celebration of the Music from Coco",
           "Atlantis Rising", "Bizarre Dinosaurs",
           "Breaking2", "El Materdor", "Spinning",
           "Unidentified Flying Mater", "Bugged",
           "Decorating Disney: Holiday Magic",
           "Jonas", "Disney Junior Music Nursery Rhymes",
           "Minnies Bow-Toons", "Under the Sea: A Descendants Story",
           "Holiday Magic", "The Adventures of Spin and Marty"]

good_reviews = [6.0,6.1,6.2,6.3,6.4,6.5,6.6,6.7,6.8,6.9,7,7.1,7.2,
               7.3,7.4,7.5,7.6,7.7,7.8,7.9,8,8.1,8.2,8.3,8.4,8.5,
               8.6,8.7,8.8,8.9,9,9.1,9.2,9.3,9.4,9.5,9.6,9.7,9.8,9.9]

bad_reviews = [1.5,2.3,3,3.3,3.4,3.5,3.6,3.7,4,4.1,4.2,4.3,4.4,4.5,
              4.6,4.7,4.8,4.9,5,5.1,5.2,5.3,5.4,5.5,5.6,5.7,5.8,5.9]

ten_list = ["Disney Junior Music Nursery Rhymes", "Minnies Bow-Toons",
            "Wind"]

twentytwo_list = ["101 Dalmatian Street", "Billy Dilleys Super-Duper Subterranean Summer",
                  "Legend of the Three Caballeros"]


thirty_list = ["Dog: Impossible", "Dr. Oakley, Yukon Vet", "Dr. T, Lone Star Vet",
               "Lego Star Wars: The Yoda Chronicles", "Marvel Rising: Initiation",
               "Marvel Future Avengers", "Marvels Hero Project", "Muppet Moments",
               "The Adventures of Spin and Marty", "The Super Hero Squad Show",
               "The World According to Jeff Goldblum"]

fourtyfour_list = ["How Dogs Got Their Shapes", "Calling Dr. Pol", "Nature Untamed",
                   "Kingdom of the White Wolf", "Man Among Cheetahs", "Primal Survivor",
                   "Supercar Superbuild", "The Flood", "The Lost Tomb of Alexander the Great",
                   "Unlikely Animal Friends", "Winged Seduction: Birds of Paradise", "Year Million"]


fourtyfive_list = ["Americas National Parks", "Rocky Mountain Animal Rescue",
                   "Sharks of Lost Island", "Shop Class"]

sixty_list = ["Expedition Amelia", "Encore", "Lost Cities with Albert Lin",
              "Lost Treasures of the Maya", "Secrets of the King Cobra",
              "Tutankhamens Treasures"]


# Filling in N/A and None cells for the language column and ID column
for a in range(2, 896):
    lang = ws.cell(row = a, column = 14).value
    eng = "English"
    
    if lang in not_available:
        ws.cell(row = a, column = 14).value = eng
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = a, column = 14).fill = corrected_cells
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = a, column = 1).fill = corrected_cells



# Filling in N/A cells for the country column
for a in range(2, 896):
    lang = ws.cell(row = a, column = 15).value
    america = "USA"
    
    if lang in not_available:
        ws.cell(row = a, column = 15).value = america
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = a, column = 15).fill = corrected_cells
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = a, column = 1).fill = corrected_cells

        


# Filling in Not Rated labels for the rated column
for b in range(2, 896):
    tile = ws.cell(row = b, column = 2).value
    nr = "Not Rated"
    
    if tile in nr_list:
        ws.cell(row = b, column = 5).value = nr
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = b, column = 5).fill = corrected_cells
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = b, column = 1).fill = corrected_cells


# Filling in Approved labels for the rated column
for c in range(2, 896):
    tile = ws.cell(row = c, column = 2).value
    appvd = "Approved"
    
    if tile in approv_list:
        ws.cell(row = c, column = 5).value = appvd
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = c, column = 5).fill = corrected_cells
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = c, column = 1).fill = corrected_cells


# Filling in G labels for the rated column
for d in range(2, 896):
    tile = ws.cell(row = d, column = 2).value
    g_rated = "G"
    
    if tile in g_list:
        ws.cell(row = d, column = 5).value = g_rated
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = d, column = 5).fill = corrected_cells
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = d, column = 1).fill = corrected_cells


# Filling in TV-G labels for the rated column
for e in range(2, 896):
    tile = ws.cell(row = e, column = 2).value
    tvg_rated = "TV-G"
    
    if tile in tvg_list:
        ws.cell(row = e, column = 5).value = tvg_rated
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = e, column = 5).fill = corrected_cells
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = e, column = 1).fill = corrected_cells
        


# Filling in TV-PG labels for the rated column
for f in range(2, 896):
    tile = ws.cell(row = f, column = 2).value
    tvpg_rated = "TV-PG"
    
    if tile in tvpg_list:
        ws.cell(row = f, column = 5).value = tvpg_rated
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = f, column = 5).fill = corrected_cells
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = f, column = 1).fill = corrected_cells


# Filling in TV-Y labels for the rated column
for g in range(2, 896):
    tile = ws.cell(row = g, column = 2).value
    tvy_rated = "TV-Y"
    
    if tile in tvy_list:
        ws.cell(row = g, column = 5).value = tvy_rated
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = g, column = 5).fill = corrected_cells
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = g, column = 1).fill = corrected_cells


# Filling in TV-Y7 labels for the rated column
for h in range(2, 896):
    tile = ws.cell(row = h, column = 2).value
    tvy7_rated = "TV-Y7"
    
    if tile in tvy7_list:
        ws.cell(row = h, column = 5).value = tvy7_rated
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = h, column = 5).fill = corrected_cells
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = h, column = 1).fill = corrected_cells


# Filling in TV-14 labels for the rated column
for i in range(2, 896):
    tile = ws.cell(row = i, column = 2).value
    tv14_rated = "TV-14"
    
    if tile in tv14_list:
        ws.cell(row = i, column = 5).value = tv14_rated
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = i, column = 5).fill = corrected_cells
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = i, column = 1).fill = corrected_cells


# Filling in PG labels for the rated column
for j in range(2, 896):
    tile = ws.cell(row = j, column = 2).value
    pg_rated = "PG"
    
    if tile in pg_list:
        ws.cell(row = j, column = 5).value = pg_rated
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = j, column = 5).fill = corrected_cells
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = j, column = 1).fill = corrected_cells



# Replace names with the wrong accent marks in cells
ws['M35'] = "Jackie Chan, Steve Coogan, Cécile de France, Robert Fyfe"
ws['M35'].fill = corrected_cells
ws['A35'].fill = corrected_cells

ws['L36'] = "Georgeos Díaz-Montexano (story advisor), Simcha Jacobovici"
ws['L36'].fill = corrected_cells

ws['M36'] = "James Cameron, Georgeos Díaz-Montexano, Peter Ellul Vincenti, Richard Freund"
ws['M36'].fill = corrected_cells

ws['M57'] = "Leonardo DiCaprio, Ki-moon Ban, Alejandro G. Iñárritu, Mike Brune"
ws['M57'].fill = corrected_cells
ws['A57'].fill = corrected_cells

ws['M113'] = "Anthony Gonzalez, Gael García Bernal, Benjamin Bratt, Alanna Ubach"
ws['M113'].fill = corrected_cells
ws['A113'].fill = corrected_cells

ws['M259'] = "Cozi Zuehlsdorff, Heidi Blickenstaff, Jason Maybaum, Alex Désert"
ws['M259'].fill = corrected_cells
ws['A259'].fill = corrected_cells

ws['M497'] = "Angelica Bolognesi Bonacini, Kimberlea Berg, Jim Cummings, Chloë Grace Moretz"
ws['M497'].fill = corrected_cells
ws['A497'].fill = corrected_cells

ws['M542'] = "Johnny Depp, Penélope Cruz, Geoffrey Rush, Ian McShane"
ws['M542'].fill = corrected_cells
ws['A542'].fill = corrected_cells


ws['K545'] = "Mark A.Z. Dippé"
ws['K545'].fill = corrected_cells
ws['A545'].fill = corrected_cells

ws['M588'] = "Fred Shields, José Oliveira"
ws['M588'].fill = corrected_cells
ws['A588'].fill = corrected_cells

ws['M607'] = "Stian Smestad, Gabriel Byrne, Louisa Milwood-Haigh, Trond Peter Stamsø Munch"
ws['M607'].fill = corrected_cells
ws['A607'].fill = corrected_cells

ws['M618'] = "Cuba Gooding Jr., James Coburn, Sisqó, Nichelle Nichols"
ws['M618'].fill = corrected_cells
ws['A618'].fill = corrected_cells

ws['M645'] = "Grace VanderWaal, Graham Verchere, Giancarlo Esposito, Maximiliano Hernández"
ws['M645'].fill = corrected_cells
ws['A645'].fill = corrected_cells

ws['M727'] = "Émile Genest, John Drainie, Tommy Tweed, Sandra Scott"
ws['M727'].fill = corrected_cells
ws['A727'].fill = corrected_cells

ws['M816'] = "Richard Romanus, Biana Tamimi, Patrick Elyas, Gérard Rudolf"
ws['M816'].fill = corrected_cells
ws['A816'].fill = corrected_cells

ws['M885'] = "Laurence Fishburne, Miklós Bányai, Joe Corrigall, Siobhan Dillon"
ws['M885'].fill = corrected_cells



# Filling empty cells for runtime
ws['I10'] = "47 min"
ws['I10'].fill = corrected_cells

ws['I47'] = "25 min"
ws['I47'].fill = corrected_cells

ws['I64'] = "48 min"
ws['I64'].fill = corrected_cells

ws['I67'] = "7 min"
ws['I67'].fill = corrected_cells
ws['A67'].fill = corrected_cells

ws['I147'] = "8 min"
ws['I147'].fill = corrected_cells

ws['I153'] = "17 min"
ws['I153'].fill = corrected_cells

ws['I223'] = "120 min"
ws['I223'].fill = corrected_cells

ws['I228'] = "86 min"
ws['I228'].fill = corrected_cells

ws['I231'] = "95 min"
ws['I231'].fill = corrected_cells

ws['I367'] = "94 min"
ws['I367'].fill = corrected_cells

ws['I378'] = "6 min"
ws['I378'].fill = corrected_cells
ws['A378'].fill = corrected_cells

ws['I452'] = "2 min"
ws['I452'].fill = corrected_cells

ws['I521'] = "43 min"
ws['I521'].fill = corrected_cells

ws['I522'] = "77 min"
ws['I522'].fill = corrected_cells

ws['I536'] = "80 min"
ws['I536'].fill = corrected_cells

ws['I544'] = "4 min"
ws['I544'].fill = corrected_cells
ws['A544'].fill = corrected_cells

ws['I599'] = "46 min"
ws['I599'].fill = corrected_cells

ws['I693'] = "89 min"
ws['I693'].fill = corrected_cells



# Filling empty cells for genre
ws['J47'] = "Game-Show, Reality-TV"
ws['J47'].fill = corrected_cells

ws['J156'] = "Animation, Family, Music"
ws['J156'].fill = corrected_cells

ws['J176'] = "Reality-TV"
ws['J176'].fill = corrected_cells

ws['J217'] = "Documentary"
ws['J217'].fill = corrected_cells

ws['J235'] = "Documentary"
ws['J235'].fill = corrected_cells

ws['J522'] = "Documentary"
ws['J522'].fill = corrected_cells

ws['J604'] = "Documentary"
ws['J604'].fill = corrected_cells

ws['J608'] = "Adventure,Family,Game-Show"
ws['J608'].fill = corrected_cells

ws['J882'] = "Reality-TV"
ws['J882'].fill = corrected_cells



# Filling empty cells for imdb rating and votes
ws['R47'] = 6.5
ws['R47'].fill = corrected_cells

ws['S47'] = 181
ws['S47'].fill = corrected_cells

ws['R156'] = 6.4
ws['R156'].fill = corrected_cells

ws['S156'] = 23
ws['S156'].fill = corrected_cells

ws['R176'] = 6.8
ws['R176'].fill = corrected_cells

ws['S176'] = 136
ws['S176'].fill = corrected_cells

ws['R185'] = 7.4
ws['R185'].fill = corrected_cells

ws['S185'] = 114
ws['S185'].fill = corrected_cells

ws['R199'] = 7.8
ws['R199'].fill = corrected_cells

ws['S199'] = 132
ws['S199'].fill = corrected_cells

ws['R335'] = 8.5
ws['R335'].fill = corrected_cells

ws['S335'] = 1800
ws['S335'].fill = corrected_cells

ws['R351'] = 8.5
ws['R351'].fill = corrected_cells

ws['S351'] = 1800
ws['S351'].fill = corrected_cells

ws['R356'] = 6.6
ws['R356'].fill = corrected_cells

ws['S356'] = 16
ws['S356'].fill = corrected_cells

ws['R420'] = 5.1
ws['R420'].fill = corrected_cells
ws['A420'].fill = corrected_cells

ws['S420'] = 398
ws['S420'].fill = corrected_cells

ws['R536'] = 7.6
ws['R536'].fill = corrected_cells

ws['S536'] = 1500
ws['S536'].fill = corrected_cells

ws['R595'] = 7.2
ws['R595'].fill = corrected_cells

ws['S595'] = 60
ws['S595'].fill = corrected_cells

ws['R604'] = 5.6
ws['R604'].fill = corrected_cells

ws['S604'] = 134
ws['S604'].fill = corrected_cells

ws['R726'] = 8.0
ws['R726'].fill = corrected_cells

ws['S726'] = 8
ws['S726'].fill = corrected_cells

ws['R873'] = 8.1
ws['R873'].fill = corrected_cells

ws['S873'] = 14
ws['S873'].fill = corrected_cells

ws['R882'] = 0
ws['R882'].fill = corrected_cells

ws['S882'] = 0
ws['S882'].fill = corrected_cells


# Filling empty cells for 101 Dalmantian Street
ws['K3'] = "Miklós Weigert, Jez Hall, Frédéric Martin, Joonas Utti"
ws['K3'].fill = corrected_cells
ws['A3'].fill = corrected_cells

ws['L3'] = "Jess Kedward, Suzanne Lang, Ciaran Morrison, Mick O'Hara, Maria O'Loughlin"
ws['L3'].fill = corrected_cells

# Filling empty cells for a celecbration of the music from coco
ws['C10'] = "A concert style performance at the Hollywood Bowl with some of the cast of Coco and singers and dancers. A vibrant celebration of culture, love, family, and music."
ws['C10'].fill = corrected_cells

ws['L10'] = "Brittany Thompson, Michelle Zagorsky"
ws['L10'].fill = corrected_cells

# Filling empty cells for americas funniest home videos
ws['K27'] = "Vin Di Bona, Steve Hirsen,nE.C. Pauling, Averill Perry, Rob Katz, Russ Reinsel, Robin Felsen Von Halle"
ws['K27'].fill = corrected_cells
ws['A27'].fill = corrected_cells

ws['L27'] = "Todd Thicke, Michael Palleschi, Erik Lohla, Bob Arnott, Bob Saget, Jordan Schatz, Trace Beaulieu, J. Elvis Weinstein"
ws['L27'].fill = corrected_cells

# Filling empty cells for Be our chef row
ws['C47'] = "A cooking competition that challenges five food-loving families to create delicious dishes inspired by the magic of Disney. In each episode, two families go head-to-head in a themed cooking challenge at Walt Disney World."
ws['C47'].fill = corrected_cells

ws['K47'] = "Adam Vetri"
ws['K47'].fill = corrected_cells

ws['M47'] = "Angela Kinsey, DeMaryo Platt"
ws['M47'].fill = corrected_cells

# Filling empty cells for Lengend of three caballeros
ws['C159'] = "Donald Duck, Jóse Carioca and Panchito Pistoles find themselves inheritors of a heroic legacy, and resume their ancestors' quest with the aid of a valiant Greek goddess."
ws['C159'].fill = corrected_cells
ws['A159'].fill = corrected_cells

# Filling empty cells for Soy Luna
ws['C625'] = "Luna Valente lives with her family in Cancún, Mexico. She goes to school, has her own group of friends, has a job, and loves to skate. However, her life changes when her parents are given a job offer that moves them to Buenos Aires, Argentina. There she finds a skating rink named Jam and Roller where she learns free styling. She makes new friends and falls in love with Matteo Balsano."
ws['C625'].fill = corrected_cells

# Filling empty cells for Pegunis
ws['C192'] = "A coming-of-age story about an Adélie penguin named Steve who joins millions of fellow males in the icy Antarctic spring on a quest to build a suitable nest, find a life partner and start a family."
ws['C192'].fill = corrected_cells
ws['A192'].fill = corrected_cells

# Filling empty cells for My Friends Tigger & Pooh
ws['K459'] = "David Hartman, Don MacKinnon"
ws['K459'].fill = corrected_cells
ws['A459'].fill = corrected_cells

ws['L459'] = "Kim Beyer-Johnson, Eileen Cabiling, Nicole Dubuc, Erika Grediaga, Brian Hohlfeld, Ron Holsey, Janna King, Catherine Lieuwen"
ws['L459'].fill = corrected_cells

# Filling empty cells for Thor:Ragnarok
ws['C440'] = "Imprisoned on the planet Sakaar, Thor must race against time to return to Asgard and stop Ragnarök, the destruction of his world, at the hands of the powerful and ruthless villain Hela."
ws['C440'].fill = corrected_cells
ws['A440'].fill = corrected_cells


# Filling empty cells for Year Million
ws['K885'] = "Mark Elijah Rosenberg, Walter Pitt"
ws['K885'].fill = corrected_cells

ws['L885'] = "Chris Connolly, Jenny Connell Davis, Wendy Greene, Jeremy Lubman, Bryan Wizemann"
ws['L885'].fill = corrected_cells


# Update the Good Review column with yes or no 
for k in range(2, 896):
    rating = ws.cell(row = k, column = 18).value
    yea = "Y"
    nope = "N"

    if rating in good_reviews:
        ws.cell(row = k, column = 20).value = yea

    else: 
        ws.cell(row = k, column = 20).value = nope


# Update the Bad Review column with yes or no 
for l in range(2, 896):
    rating = ws.cell(row = l, column = 18).value
    yea = "Y"
    nope = "N"

    if rating in bad_reviews:
        ws.cell(row = l, column = 21).value = yea

    else: 
        ws.cell(row = l, column = 21).value = nope

        

# Filling in 10 mins labels for the runtime column
for m in range(2, 896):
    rtime = ws.cell(row = m, column = 9).value
    ten_min = "10 min"
    
    if rtime in ten_list:
        ws.cell(row = h, column = 9).value = ten_min
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = h, column = 9).fill = corrected_cells
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = h, column = 1).fill = corrected_cells


# Filling in 22 mins labels for the runtime column
for n in range(2, 896):
    rtime = ws.cell(row = n, column = 9).value
    twentytwo_min = "22 min"
    
    if rtime in twentytwo_list:
        ws.cell(row = n, column = 9).value = twentytwo_min
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = n, column = 9).fill = corrected_cells
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = n, column = 1).fill = corrected_cells


# Filling in 30 mins labels for the runtime column
for o in range(2, 896):
    rtime = ws.cell(row = o, column = 9).value
    thirty_min = "30 min"
    
    if rtime in thirty_list:
        ws.cell(row = o, column = 9).value = thirty_min
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = o, column = 9).fill = corrected_cells
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = o, column = 1).fill = corrected_cells


# Filling in 44 mins labels for the runtime column
for p in range(2, 896):
    rtime = ws.cell(row = p, column = 9).value
    fourtyfour_min = "44 min"
    
    if rtime in fourtyfour_list:
        ws.cell(row = p, column = 9).value = fourtyfour_min
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = p, column = 9).fill = corrected_cells
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = p, column = 1).fill = corrected_cells


# Filling in 45 mins labels for the runtime column
for q in range(2, 896):
    rtime = ws.cell(row = q, column = 9).value
    fourtyfive_min = "45 min"
    
    if rtime in fourtyfive_list:
        ws.cell(row = q, column = 9).value = fourtyfive_min
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = q, column = 9).fill = corrected_cells
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = q, column = 1).fill = corrected_cells


# Filling in 60 mins labels for the runtime column
for r in range(2, 896):
    rtime = ws.cell(row = r, column = 9).value
    sixty_min = "60 min"
    
    if rtime in sixty_list:
        ws.cell(row = r, column = 9).value = sixty_min
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = r, column = 9).fill = corrected_cells
        
        # Highlight the updated cells with yellow fill color
        ws.cell(row = r, column = 1).fill = corrected_cells


for s in wb.worksheets:
    # Use the iterate rows function to look for phrase in whole workbook
    for row in s.iter_rows():
        for cell in row:
            if cell.value in replacement_pair.keys():
                cell.value = replacement_pair.get(cell.value)




wb.save("disney_plus_labeling_tool_demo_v2.xlsx")



